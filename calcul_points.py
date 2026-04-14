import openpyxl
import re
import requests
import json
import time

# --- MAPPING DES ÉQUIPES ---
# Permet de faire le lien entre vos noms raccourcis et les noms officiels de la NBA
TEAM_ALIASES = {
    "cavs": "Cavaliers",
    "sixers": "76ers",
    "76ers": "76ers",
    "wolves": "Timberwolves",
    "blazers": "Trail Blazers",
    "mavs": "Mavericks",
    "spurs": "Spurs",
    "rockets": "Rockets",
    "thunder": "Thunder",
    "celtics": "Celtics",
    "knicks": "Knicks",
    "hawks": "Hawks",
    "raptors": "Raptors",
    "pistons": "Pistons",
    "magic": "Magic",
    "heat": "Heat",
    "hornets": "Hornets",
    "suns": "Suns",
    "clippers": "Clippers",
    "warriors": "Warriors",
    "lakers": "Lakers",
    "nuggets": "Nuggets",
    "pelicans": "Pelicans",
    "grizzlies": "Grizzlies",
    "kings": "Kings",
    "pacers": "Pacers",
    "bulls": "Bulls",
    "bucks": "Bucks",
    "nets": "Nets"
}

def get_nba_bracket_results():
    """
    Récupère les résultats actuels des séries depuis l'API de statistiques NBA.
    """
    url = "https://stats.nba.com/stats/playoffbracket?LeagueID=00&SeasonYear=2025&State=2"
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://www.nba.com/"
    }
    series_results = {}
    try:
        response = requests.get(url, headers=headers, timeout=10)
        data = response.json()
        if 'resultSets' in data and len(data['resultSets']) > 0:
            headers_list = data['resultSets'][0]['headers']
            rows = data['resultSets'][0]['rowSet']
            for row in rows:
                series_data = dict(zip(headers_list, row))
                high_wins = series_data.get('HIGH_SEED_SERIES_WINS', 0)
                low_wins = series_data.get('LOW_SEED_SERIES_WINS', 0)
                high_team = series_data.get('HIGH_SEED_TEAM_NICKNAME', '')
                low_team = series_data.get('LOW_SEED_TEAM_NICKNAME', '')
                
                if high_wins == 4:
                    series_results[high_team.lower()] = 4 + low_wins
                elif low_wins == 4:
                    series_results[low_team.lower()] = 4 + high_wins
    except Exception as e:
        print(f"Erreur lors de la récupération des données: {e}")
        
    return series_results

def get_mvps():
    """
    Les MVP ne sont pas facilement récupérables via le bracket NBA.
    En fin de saison, vous pourrez venir les inscrire ici.
    """
    return {
        "MVP de l’EST": "tatum",  # Exemple: Remplacer par le vrai joueur
        "MVP de l’OUEST": "jokic", 
        "MVP des FINALES": "tatum"
    }


def parse_prediction(text):
    """
    Transforme "Pistons en 6" en ("Pistons", 6)
    """
    if not text or not isinstance(text, str):
        return None, None
        
    text = str(text).strip()
    # On cherche le mot clé de l'équipe (tout avant le chiffre) et le chiffre
    match = re.search(r'([A-Za-z\s]+)(?:en\s)?(\d)', text, re.IGNORECASE)
    if match:
        team_str = match.group(1).strip().lower()
        games = int(match.group(2))
        
        # Résoudre l'alias (ex: 'cavs' -> 'cavaliers')
        team_name = team_str
        for alias, official in TEAM_ALIASES.items():
            if alias in team_str:
                team_name = official.lower()
                break
                
        return team_name, games
    
    # Si format différent (juste une équipe sans chiffre ?)
    return text.lower(), None

def calculer_points(fichier_excel):
    print("Chargement du fichier Excel...")
    wb = openpyxl.load_workbook(fichier_excel)
    ws = wb['Sheet1']
    
    # Récupérer les vrais résultats (Si l'API est vide, on peut temporairement ajouter un "Mock" pour tester)
    # mock_results = {"pistons": 6, "cavaliers": 5, "knicks": 7}
    print("Récupération des résultats sur le site de la NBA...")
    real_results = get_nba_bracket_results()
    real_mvps = get_mvps()
    
    print(f"Séries terminées trouvées : {real_results}")

    # Les lignes de séries dans l'Excel
    lignes_series = list(range(21, 29)) + list(range(30, 34)) + [35, 37, 40]
    # Lignes pour les MVPs
    lignes_mvps = [36, 38, 41]
    
    # Les colonnes de prédictions (WILL, OLI, MIK, THOM, SAM, DOUSKI)
    colonnes_joueurs = [
        {'pred': 'C', 'score': 'D'},
        {'pred': 'E', 'score': 'F'},
        {'pred': 'G', 'score': 'H'},
        {'pred': 'I', 'score': 'J'},
        {'pred': 'K', 'score': 'L'},
        {'pred': 'M', 'score': 'N'},
    ]

    for ligne in lignes_series:
        matchup = ws[f'B{ligne}'].value
        if not matchup:
            continue
            
        for joueur in colonnes_joueurs:
            pred_cell = ws[f"{joueur['pred']}{ligne}"].value
            score_cell = f"{joueur['score']}{ligne}"
            
            if pred_cell:
                team_pred, games_pred = parse_prediction(pred_cell)
                if team_pred:
                    # On check si l'équipe match les résultats réels
                    points = 0
                    if team_pred in real_results:
                        real_games = real_results[team_pred]
                        # A la bonne équipe qui a gagné la série = 1pt
                        points += 1
                        # A aussi le bon nombre de matchs = 2pts
                        if games_pred == real_games:
                            points += 1
                        
                        ws[score_cell] = points

    # Attribution des points pour MVP
    for ligne in lignes_mvps:
        titre_mvp = ws[f'B{ligne}'].value
        if titre_mvp and titre_mvp in real_mvps:
            vrai_mvp = real_mvps[titre_mvp].lower().strip()
            for joueur in colonnes_joueurs:
                pred_cell = ws[f"{joueur['pred']}{ligne}"].value
                score_cell = f"{joueur['score']}{ligne}"
                if pred_cell and str(pred_cell).lower().strip() == vrai_mvp:
                    ws[score_cell] = 1
                elif pred_cell:
                    ws[score_cell] = 0

    # Sauvegarder
    wb.save(fichier_excel)
    print("Points mis à jour avec succès dans le fichier !")

if __name__ == "__main__":
    calculer_points("Prédictions NBA 2026.xlsx")
